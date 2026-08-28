#!/usr/bin/env python3
"""
Genera PokemonGO_Tracker.xlsx — tracker manuale con calcoli automatici.
Celle gialle = input manuale | Celle verdi = formule
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "PokemonGO_Tracker.xlsx"

# Colori
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="C6EFCE")
HEADER = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

GEN_TOTALS = {1: 151, 2: 100, 3: 135, 4: 107, 5: 156, 6: 72, 7: 88, 8: 96, 9: 120}

# Tabella XP cumulativo (aggiornamento livelli 1–80, ottobre 2025)
XP_TABLE = [
    0, 2500, 5500, 9000, 13000, 18000, 24000, 31000, 39000, 48000,
    58000, 70000, 84000, 100000, 118000, 139000, 163500, 191500, 223000, 258000,
    300000, 349000, 405000, 468000, 538000, 621000, 717000, 826000, 948000, 1083000,
    1241000, 1422000, 1626000, 1853000, 2103000, 2393000, 2723000, 3093000, 3503000, 3953000,
    4473000, 5063000, 5723000, 6453000, 7253000, 8153000, 9153000, 10253000, 11453000, 12753000,
    14193000, 15773000, 17493000, 19353000, 21353000, 23553000, 25953000, 28553000, 31353000, 34353000,
    37703000, 41403000, 45453000, 49853000, 54603000, 59853000, 65603000, 71853000, 78603000, 85853000,
    93853000, 102603000, 112103000, 122353000, 133353000, 145353000, 158353000, 172353000, 187353000, 203353000,
]

MAX_LEVEL = 80
XP_SHEET_END_ROW = 3 + MAX_LEVEL  # riga 83 per livello 80

MEDALS_PATH = Path(__file__).parent / "medals.json"
POKEMON_PATH = Path(__file__).parent / "pokemon.json"


def load_medals():
    with open(MEDALS_PATH, encoding="utf-8") as f:
        return json.load(f)


def load_pokemon():
    with open(POKEMON_PATH, encoding="utf-8") as f:
        return json.load(f)


def write_medals_js(medals):
    out = Path(__file__).parent / "app" / "medals.js"
    out.write_text(
        "// Generato da generate_excel.py — non modificare manualmente\n"
        f"const MEDALS = {json.dumps(medals, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )


def write_pokemon_js(pokemon):
    out = Path(__file__).parent / "app" / "pokemon.js"
    out.write_text(
        "// Generato da generate_excel.py — non modificare manualmente\n"
        f"const POKEMON = {json.dumps(pokemon, ensure_ascii=False)};\n",
        encoding="utf-8",
    )


MEDALS = load_medals()
POKEMON = load_pokemon()

LEGENDARIES = [
    "Articuno", "Zapdos", "Moltres", "Mewtwo", "Raikou", "Entei", "Suicune",
    "Lugia", "Ho-Oh", "Regirock", "Regice", "Registeel", "Latias", "Latios",
    "Kyogre", "Groudon", "Rayquaza", "Uxie", "Mesprit", "Azelf", "Dialga",
    "Palkia", "Heatran", "Regigigas", "Giratina", "Cresselia", "Cobalion",
    "Terrakion", "Virizion", "Tornadus", "Thundurus", "Reshiram", "Zekrom",
    "Landorus", "Kyurem", "Xerneas", "Yveltal", "Zygarde", "Tapu Koko",
    "Tapu Lele", "Tapu Bulu", "Tapu Fini", "Solgaleo", "Lunala", "Necrozma",
    "Zacian", "Zamazenta", "Eternatus", "Regieleki", "Regidrago", "Glastrier",
    "Spectrier", "Calyrex", "Koraidon", "Miraidon",
]


def style_cell(cell, fill=None, font=None, alignment=None, border=True, number_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = BORDER
    if number_format:
        cell.number_format = number_format


def set_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        style_cell(c, HEADER, HEADER_FONT, CENTER)


def build_xp_sheet(wb):
    ws = wb.create_sheet("XP")
    ws["A1"] = "Tabella Livelli e XP Cumulativo (1–80)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Nota: i livelli 71–80 richiedono anche ricerche di salita di livello oltre all'XP."
    set_header_row(ws, 3, ["Livello", "XP Cumulativo"])
    for lvl in range(1, MAX_LEVEL + 1):
        r = lvl + 3
        ws.cell(r, 1, lvl)
        ws.cell(r, 2, XP_TABLE[lvl - 1])
        for col in (1, 2):
            style_cell(ws.cell(r, col), GREEN, alignment=CENTER, number_format="#,##0" if col == 2 else "0")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18


def build_pokedex_sheet(wb):
    ws = wb.create_sheet("Pokedex")
    ws["A1"] = "Pokédex per Generazione"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Inserisci manualmente Catturati e Visti (celle gialle)."
    set_header_row(ws, 4, ["Gen", "Totale", "Catturati", "Visti", "Mancanti", "% Catturati", "% Visti"])
    for i, gen in enumerate(range(1, 10)):
        r = 5 + i
        ws.cell(r, 1, f"Gen {gen}")
        ws.cell(r, 2, GEN_TOTALS[gen])
        ws.cell(r, 3, 0)  # Catturati - manual
        ws.cell(r, 4, 0)  # Visti - manual
        ws.cell(r, 5, f"=B{r}-C{r}")
        ws.cell(r, 6, f"=IF(B{r}=0,0,C{r}/B{r})")
        ws.cell(r, 7, f"=IF(B{r}=0,0,D{r}/B{r})")
        style_cell(ws.cell(r, 1), alignment=CENTER)
        style_cell(ws.cell(r, 2), GREEN, alignment=CENTER)
        style_cell(ws.cell(r, 3), YELLOW, alignment=CENTER)
        style_cell(ws.cell(r, 4), YELLOW, alignment=CENTER)
        for col in (5, 6, 7):
            fmt = "0.00%" if col >= 6 else "0"
            style_cell(ws.cell(r, col), GREEN, alignment=CENTER, number_format=fmt)
    # Riga totale
    r = 14
    ws.cell(r, 1, "TOTALE")
    ws.cell(r, 1).font = BOLD
    ws.cell(r, 2, f"=SUM(B5:B13)")
    ws.cell(r, 3, f"=SUM(C5:C13)")
    ws.cell(r, 4, f"=SUM(D5:D13)")
    ws.cell(r, 5, f"=B{r}-C{r}")
    ws.cell(r, 6, f"=IF(B{r}=0,0,C{r}/B{r})")
    ws.cell(r, 7, f"=IF(B{r}=0,0,D{r}/D{r})" if False else f"=IF(B{r}=0,0,D{r}/B{r})")
    for col in range(1, 8):
        fmt = "0.00%" if col >= 6 else "0"
        fill = GREEN
        style_cell(ws.cell(r, col), fill, BOLD, CENTER, number_format=fmt if col >= 6 or col == 2 else "0")
    for col, w in zip("ABCDEFG", [8, 10, 12, 12, 12, 14, 14]):
        ws.column_dimensions[col].width = w


def build_shiny_sheet(wb):
    ws = wb.create_sheet("Shiny")
    ws["A1"] = "Registro Shiny"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Compila le righe gialle. IV% si calcola automaticamente."
    set_header_row(ws, 4, ["Data", "Pokémon", "CP", "IV Att", "IV Dif", "IV PS", "IV %", "Metodo", "Note"])
    for r in range(5, 55):
        for col in range(1, 10):
            cell = ws.cell(r, col)
            if col in (1, 2, 3, 4, 5, 6, 8, 9):
                style_cell(cell, YELLOW, alignment=CENTER if col != 9 else LEFT)
            elif col == 7:
                cell.value = f'=IF(OR(D{r}="",E{r}="",F{r}=""),"",ROUND((D{r}+E{r}+F{r})/45*100,2))'
                style_cell(cell, GREEN, alignment=CENTER, number_format='0.00"%"')
    widths = [12, 18, 8, 8, 8, 8, 10, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_medals_sheet(wb):
    ws = wb.create_sheet("Medaglie")
    ws["A1"] = "Progresso Medaglie"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Inserisci il progresso attuale (colonna D). Tier e % verso il prossimo livello sono automatici."
    set_header_row(ws, 4, [
        "Medaglia", "Categoria", "Descrizione", "Progresso",
        "Bronzo", "Argento", "Oro", "Platino", "Tier", "% Prossimo",
    ])
    start = 5
    for i, m in enumerate(MEDALS):
        r = start + i
        ws.cell(r, 1, m["name"])
        ws.cell(r, 2, m["category"])
        ws.cell(r, 3, m["desc"])
        ws.cell(r, 4, 0)
        ws.cell(r, 5, m["bronze"])
        ws.cell(r, 6, m["silver"])
        ws.cell(r, 7, m["gold"])
        ws.cell(r, 8, m["platinum"])
        tier = (
            f'=IF(D{r}>=H{r},"Platino",'
            f'IF(D{r}>=G{r},"Oro",'
            f'IF(D{r}>=F{r},"Argento",'
            f'IF(D{r}>=E{r},"Bronzo","Nessuno"))))'
        )
        pct = (
            f'=IF(D{r}>=H{r},1,'
            f'IF(D{r}>=G{r},IF(H{r}=G{r},1,(D{r}-G{r})/(H{r}-G{r})),'
            f'IF(D{r}>=F{r},IF(G{r}=F{r},1,(D{r}-F{r})/(G{r}-F{r})),'
            f'IF(D{r}>=E{r},IF(F{r}=E{r},1,(D{r}-E{r})/(F{r}-E{r})),'
            f'IF(E{r}=0,0,D{r}/E{r})))))'
        )
        ws.cell(r, 9, tier)
        ws.cell(r, 10, pct)
        style_cell(ws.cell(r, 1), alignment=LEFT)
        style_cell(ws.cell(r, 2), alignment=LEFT)
        style_cell(ws.cell(r, 3), alignment=LEFT)
        style_cell(ws.cell(r, 4), YELLOW, alignment=CENTER, number_format="#,##0.##")
        for col in (5, 6, 7, 8):
            style_cell(ws.cell(r, col), GREEN, alignment=CENTER, number_format="#,##0.##")
        style_cell(ws.cell(r, 9), GREEN, alignment=CENTER)
        style_cell(ws.cell(r, 10), GREEN, alignment=CENTER, number_format="0.00%")
    for col, w in zip("ABCDEFGHIJ", [18, 14, 32, 12, 10, 10, 10, 10, 12, 12]):
        ws.column_dimensions[col].width = w


def build_species_sheet(wb):
    ws = wb.create_sheet("Specie")
    ws["A1"] = "Pokédex per Specie (1025)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Segna Visto/Catturato/Shiny. Colonne Ottiene/Evoluzione sono di riferimento."
    set_header_row(ws, 4, ["#", "Nome", "Gen", "Tipi", "Boost meteo", "Debolezze", "Resistenze", "Ottiene", "Evoluzione", "Visto", "Catturato", "Shiny"])
    start = 5
    for i, p in enumerate(POKEMON):
        r = start + i
        obtain = "; ".join(p.get("obtain", []))
        evo_parts = []
        for e in p.get("evolvesTo", []):
            evo_parts.append(f"→ {e['name']}: {e['how']}")
        for e in p.get("evolvesFrom", []):
            evo_parts.append(f"← {e['fromName']}: {e['how']}")
        evo = " | ".join(evo_parts) if evo_parts else "—"
        weather = "; ".join(
            f"{w['weather']} ({', '.join(w['types'])})" for w in p.get("weatherBoosts", [])
        )
        wk = p.get("weaknesses", {})
        weak = ", ".join([*(f"{t} x2.56" for t in wk.get("doubleWeak", [])), *(f"{t} x1.6" for t in wk.get("weak", []))])
        resist = ", ".join([*(f"{t} x0.39" for t in wk.get("doubleResist", [])), *(f"{t} x0.62" for t in wk.get("resist", []))])
        ws.cell(r, 1, p["id"])
        ws.cell(r, 2, p["name"])
        ws.cell(r, 3, p["gen"])
        ws.cell(r, 4, ", ".join(p.get("types", [])))
        ws.cell(r, 5, weather or "—")
        ws.cell(r, 6, weak or "—")
        ws.cell(r, 7, resist or "—")
        ws.cell(r, 8, obtain)
        ws.cell(r, 9, evo)
        ws.cell(r, 10, "No")
        ws.cell(r, 11, "No")
        ws.cell(r, 12, "No")
        style_cell(ws.cell(r, 1), GREEN, alignment=CENTER)
        style_cell(ws.cell(r, 2), alignment=LEFT)
        style_cell(ws.cell(r, 3), GREEN, alignment=CENTER)
        style_cell(ws.cell(r, 4), GREEN, alignment=LEFT)
        for col in (5, 6, 7, 8, 9):
            style_cell(ws.cell(r, col), GREEN, alignment=LEFT)
        for col in (10, 11, 12):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER)
    for col, w in zip("ABCDEFGHIJKL", [8, 18, 6, 14, 28, 28, 28, 36, 36, 10, 12, 10]):
        ws.column_dimensions[col].width = w


def build_events_sheet(wb):
    ws = wb.create_sheet("Eventi")
    ws["A1"] = "Registro Eventi"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 4, ["Data", "Nome", "Tipo", "Pokémon", "Shiny ottenuti", "Note"])
    for r in range(5, 55):
        for col in range(1, 7):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER if col != 6 else LEFT)
    widths = [12, 22, 16, 16, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_quests_sheet(wb):
    ws = wb.create_sheet("Ricerche")
    ws["A1"] = "Ricerche Speciali e sul Campo"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 4, ["Nome", "Tipo", "Stato", "Data inizio", "Data completamento", "Note"])
    for r in range(5, 55):
        for col in range(1, 7):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER if col != 6 else LEFT)
    widths = [24, 14, 12, 14, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_buddies_sheet(wb):
    ws = wb.create_sheet("Buddy")
    ws["A1"] = "Registro Buddy"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 4, ["Pokémon", "Attivo", "Amicizia", "Km", "Caramelle", "Cuori", "Data inizio"])
    for r in range(5, 55):
        for col in range(1, 8):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER)
    widths = [18, 10, 16, 10, 12, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_showcase_sheet(wb):
    ws = wb.create_sheet("Vetrina")
    ws["A1"] = "Vetrina — Migliori Catture"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 4, ["Data", "Pokémon", "CP", "IV Att", "IV Dif", "IV PS", "IV %", "Tag"])
    for r in range(5, 55):
        for col in range(1, 9):
            if col == 7:
                ws.cell(r, 7, f'=IF(OR(D{r}="",E{r}="",F{r}=""),"",ROUND((D{r}+E{r}+F{r})/45*100,2))')
                style_cell(ws.cell(r, col), GREEN, alignment=CENTER, number_format='0.00"%"')
            else:
                style_cell(ws.cell(r, col), YELLOW, alignment=CENTER)
    widths = [12, 18, 8, 8, 8, 8, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_battles_sheet(wb):
    ws = wb.create_sheet("Battaglie")
    ws["A1"] = "Battaglie e Buddy"
    ws["A1"].font = TITLE_FONT

    # Raid
    ws["A3"] = "RAID"
    ws["A3"].font = BOLD
    labels = [("Vittorie", "B4", YELLOW), ("Sconfitte", "B5", YELLOW), ("Win Rate", "B6", GREEN)]
    ws["A4"], ws["A5"], ws["A6"] = "Vittorie", "Sconfitte", "Win Rate"
    ws["B4"], ws["B5"] = 0, 0
    ws["B6"] = "=IF((B4+B5)=0,0,B4/(B4+B5))"
    style_cell(ws["A4"], alignment=LEFT)
    style_cell(ws["A5"], alignment=LEFT)
    style_cell(ws["A6"], alignment=LEFT, font=BOLD)
    style_cell(ws["B4"], YELLOW, alignment=CENTER)
    style_cell(ws["B5"], YELLOW, alignment=CENTER)
    style_cell(ws["B6"], GREEN, alignment=CENTER, number_format="0.00%")

    # GBL
    ws["A8"] = "GO BATTLE LEAGUE"
    ws["A8"].font = BOLD
    ws["A9"], ws["A10"], ws["A11"] = "Vittorie", "Sconfitte", "Win Rate"
    ws["B9"], ws["B10"] = 0, 0
    ws["B11"] = "=IF((B9+B10)=0,0,B9/(B9+B10))"
    style_cell(ws["A9"], alignment=LEFT)
    style_cell(ws["A10"], alignment=LEFT)
    style_cell(ws["A11"], alignment=LEFT, font=BOLD)
    style_cell(ws["B9"], YELLOW, alignment=CENTER)
    style_cell(ws["B10"], YELLOW, alignment=CENTER)
    style_cell(ws["B11"], GREEN, alignment=CENTER, number_format="0.00%")

    # Buddy
    ws["A13"] = "BUDDY"
    ws["A13"].font = BOLD
    ws["A14"], ws["A15"], ws["A16"], ws["A17"] = "Km percorsi", "Caramelle", "Cuori", "Caramelle / Km"
    ws["B14"], ws["B15"], ws["B16"] = 0, 0, 0
    ws["B17"] = "=IF(B14=0,0,B15/B14)"
    for row, fill in [(14, YELLOW), (15, YELLOW), (16, YELLOW), (17, GREEN)]:
        style_cell(ws[f"A{row}"], alignment=LEFT)
        fmt = "0.00" if row == 17 else "0"
        style_cell(ws[f"B{row}"], fill, alignment=CENTER, number_format=fmt)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14


def build_legendaries_sheet(wb):
    ws = wb.create_sheet("Leggendari")
    ws["A1"] = "Leggendari"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Segna Catturato/Shiny con Sì/No. Win rate = Catture/Tentativi raid."
    set_header_row(ws, 4, [
        "Pokémon", "Catturato", "Shiny", "IV %", "Data", "Tentativi Raid", "Catture", "Win Rate"
    ])
    start = 5
    for i, name in enumerate(LEGENDARIES):
        r = start + i
        ws.cell(r, 1, name)
        ws.cell(r, 2, "No")
        ws.cell(r, 3, "No")
        ws.cell(r, 4, "")
        ws.cell(r, 5, "")
        ws.cell(r, 6, 0)
        ws.cell(r, 7, 0)
        ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
        style_cell(ws.cell(r, 1), alignment=LEFT)
        for col in (2, 3, 4, 5, 6, 7):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER)
        style_cell(ws.cell(r, 8), GREEN, alignment=CENTER, number_format="0.00%")
    widths = [16, 12, 10, 10, 12, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Dashboard Pokémon GO"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Celle gialle = inserimento manuale | Celle verdi = calcoli automatici"

    # Risorse
    ws["A4"] = "RISORSE"
    ws["A4"].font = BOLD
    resources = [
        ("Stardust", 5), ("Polvere Lucente", 6), ("Poké Ball", 7),
        ("Mega Ball", 8), ("Ultra Ball", 9), ("Mega Energy", 10),
        ("Caramelle XL", 11), ("XP Totale", 12),
    ]
    for label, row in resources:
        ws.cell(row, 1, label)
        ws.cell(row, 2, 0)
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), YELLOW, alignment=CENTER, number_format="#,##0")

    # Livello da XP (riferimento foglio XP)
    ws["A14"] = "LIVELLO & XP"
    ws["A14"].font = BOLD
    ws["A15"], ws["A16"], ws["A17"] = "Livello attuale", "XP per prossimo livello", "XP mancanti"
    ws["B15"] = f'=IF(B12="","",MATCH(B12,XP!$B$4:$B${XP_SHEET_END_ROW},1))'
    ws["B16"] = f'=IF(B15>={MAX_LEVEL},0,INDEX(XP!$B$4:$B${XP_SHEET_END_ROW},B15+1))'
    ws["B17"] = f'=IF(B15>={MAX_LEVEL},0,B16-B12)'
    for row in (15, 16, 17):
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), GREEN, alignment=CENTER, number_format="#,##0")

    # Pokédex
    ws["A19"] = "POKÉDEX"
    ws["A19"].font = BOLD
    ws["A20"], ws["A21"], ws["A22"] = "Totale catturati", "% Catturati", "% Visti"
    ws["B20"] = "=Pokedex!C14"
    ws["B21"] = "=Pokedex!F14"
    ws["B22"] = "=Pokedex!G14"
    for row, fmt in [(20, "#,##0"), (21, "0.00%"), (22, "0.00%")]:
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), GREEN, alignment=CENTER, number_format=fmt)

    # Shiny
    ws["A24"] = "SHINY"
    ws["A24"].font = BOLD
    ws["A25"] = "Shiny registrati"
    ws["B25"] = '=COUNTA(Shiny!B5:B54)'
    style_cell(ws["A25"], alignment=LEFT)
    style_cell(ws["B25"], GREEN, alignment=CENTER)

    # Raid & GBL (da Battaglie)
    ws["A27"] = "RAID & GBL"
    ws["A27"].font = BOLD
    ws["A28"], ws["A29"] = "Raid Win Rate", "GBL Win Rate"
    ws["B28"] = "=Battaglie!B6"
    ws["B29"] = "=Battaglie!B11"
    for row in (28, 29):
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), GREEN, alignment=CENTER, number_format="0.00%")

    # Leggendari
    ws["A31"] = "LEGGENDARI"
    ws["A31"].font = BOLD
    ws["A32"], ws["A33"] = "Catturati", "Shiny"
    ws["B32"] = '=COUNTIF(Leggendari!B5:B58,"Sì")+COUNTIF(Leggendari!B5:B58,"Si")+COUNTIF(Leggendari!B5:B58,"sì")'
    ws["B33"] = '=COUNTIF(Leggendari!C5:C58,"Sì")+COUNTIF(Leggendari!C5:C58,"Si")+COUNTIF(Leggendari!C5:C58,"sì")'
    for row in (32, 33):
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), GREEN, alignment=CENTER)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18


def build_rocket_sheet(wb):
    ws = wb.create_sheet("Rocket")
    ws["A1"] = "Team GO Rocket"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "STATISTICHE"
    ws["A3"].font = BOLD
    for row, label in [(4, "Grunt sconfitti"), (5, "Leader sconfitti"), (6, "Giovanni sconfitti")]:
        ws.cell(row, 1, label)
        ws.cell(row, 2, 0)
        style_cell(ws.cell(row, 1), alignment=LEFT)
        style_cell(ws.cell(row, 2), YELLOW, alignment=CENTER)
    ws["A8"] = "SHADOW"
    ws["A8"].font = BOLD
    set_header_row(ws, 9, ["Pokémon", "Purificato", "IV %", "Data", "Note"])
    for r in range(10, 30):
        for col in range(1, 6):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER if col > 1 else LEFT)
    widths = [18, 12, 10, 12, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_mega_sheet(wb):
    ws = wb.create_sheet("Mega")
    ws["A1"] = "Megaevoluzione"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 3, ["Pokémon", "Energia", "Mega effettuate", "Note"])
    for r in range(4, 24):
        for col in range(1, 5):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER if col in (2, 3) else LEFT)
    widths = [18, 12, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_eggs_sheet(wb):
    ws = wb.create_sheet("Uova")
    ws["A1"] = "Uova in incubazione"
    ws["A1"].font = TITLE_FONT
    set_header_row(ws, 3, [
        "Distanza (km)", "Pokémon", "Incubatrice", "Km percorsi", "Stato", "Inizio", "Schiusa", "Note"
    ])
    for r in range(4, 24):
        for col in range(1, 9):
            style_cell(ws.cell(r, col), YELLOW, alignment=CENTER if col in (1, 3, 4, 5, 6, 7) else LEFT)
    widths = [12, 18, 12, 12, 14, 12, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    import build_pokemon_go_data
    build_pokemon_go_data.build_pokemon_data()

    wb = Workbook()
    # Rimuovi foglio default
    default = wb.active
    wb.remove(default)

    build_dashboard_sheet(wb)
    build_pokedex_sheet(wb)
    build_species_sheet(wb)
    build_shiny_sheet(wb)
    build_xp_sheet(wb)
    build_medals_sheet(wb)
    build_battles_sheet(wb)
    build_buddies_sheet(wb)
    build_events_sheet(wb)
    build_quests_sheet(wb)
    build_showcase_sheet(wb)
    build_legendaries_sheet(wb)
    build_rocket_sheet(wb)
    build_mega_sheet(wb)
    build_eggs_sheet(wb)

    order = [
        "Dashboard", "Pokedex", "Specie", "Shiny", "XP", "Medaglie",
        "Battaglie", "Buddy", "Eventi", "Ricerche", "Vetrina", "Leggendari",
        "Rocket", "Mega", "Uova",
    ]
    wb._sheets.sort(key=lambda s: order.index(s.title))

    wb.save(OUTPUT)
    write_medals_js(MEDALS)
    write_pokemon_js(POKEMON)
    print(f"Creato: {OUTPUT}")
    print(f"Aggiornato: app/medals.js ({len(MEDALS)} medaglie)")
    print(f"Aggiornato: app/pokemon.js ({len(POKEMON)} specie)")


if __name__ == "__main__":
    main()
